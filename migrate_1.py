import openpyxl
import os
import numpy

#load the PDM and Arena workbooks
#make sure the workbook is named 'PDM.xlsx', and 'Arena.xlsx'

wb_pdm = openpyxl.load_workbook('PDM.xlsx')
wb_arena = openpyxl.load_workbook('Arena.xlsx')

#identify first sheets to be used
sheet_pdm = wb_pdm['Sheet1']
sheet_arena = wb_arena['Sheet1']

#number of rows and last row in PDM and Arena
num_rows_pdm = sheet_pdm.max_row
num_rows_arena = sheet_arena.max_row

#initialize table for matching number build out
matching_table = []
master_data = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']
# arena_data = ['number', 'name', 'revision', 'phase', 'owner']
# arena_data_add = []
table_index = 0

#create array of all Arena files
# for i in range(1,num_rows_arena):
#     arena_data_add = []
#     arena_data_add.append(sheet_arena[('A'+str(i))].value)
#     arena_data_add.append(sheet_arena[('B' + str(i))].value)
#     arena_data_add.append(sheet_arena[('C' + str(i))].value)
#     arena_data_add.append(sheet_arena[('D' + str(i))].value)
#     arena_data_add.append(sheet_arena[('E' + str(i))].value)
#     arena_data=numpy.vstack((arena_data,arena_data_add))


#this creates a table of all components that match in PDM and Arena, with various name parsing
for i in range(1,num_rows_pdm):
    current_pdm = 'A' + str(i)
    current_value_ext = sheet_pdm['%s' %current_pdm].value
    current_value = os.path.splitext(current_value_ext)[0] #filename minus extension
    current_value_last = current_value[-11:] #last 11 digits of filename for incorrectly named files
    current_value_first = current_value[:11] #first 11 digits of filename for incorrectly named files

    for x in range(1,num_rows_arena):
        match_data=[]
        current_arena = 'A' + str(x)
        current_arena = sheet_arena['%s' %current_arena].value

        # look up in arena after removing extension PDM
        if current_value == current_arena:
            match_data.append(current_value_ext)#start building row of data for matches
            match_data.append(current_value)
            match_data.append(sheet_pdm[('C' + str(i))].value) # revision pdm
            match_data.append(sheet_pdm[('D' + str(i))].value) # state pdm
            match_data.append(sheet_pdm[('E' + str(i))].value) # revised by pdm
            match_data.append(sheet_arena[('C' + str(x))].value) #revision arena
            match_data.append(sheet_arena[('D' + str(x))].value) #phase state arena
            master_data = numpy.vstack((master_data,match_data))
            # matching_table.append(current_value)
            # table_index = table_index + 1

            # look up in arena using last 11 characters of PDM
        elif current_value_last == current_arena:
            match_data.append(current_value_ext)#start building row of data for matches
            match_data.append(current_value_last)
            match_data.append(sheet_pdm[('C' + str(i))].value) # revision pdm
            match_data.append(sheet_pdm[('D' + str(i))].value) # state pdm
            match_data.append(sheet_pdm[('E' + str(i))].value) # revised by pdm
            match_data.append(sheet_arena[('C' + str(x))].value) #revision arena
            match_data.append(sheet_arena[('D' + str(x))].value) #phase state arena
            master_data = numpy.vstack((master_data,match_data))
            # matching_table.append(current_value_last)
            # table_index+=1

            #look up in arena using first 11 characters of PDM
        elif current_value_first == current_arena:
            match_data.append(current_value_ext)#start building row of data for matches
            match_data.append(current_value_first)
            match_data.append(sheet_pdm[('C' + str(i))].value) # revision pdm
            match_data.append(sheet_pdm[('D' + str(i))].value) # state pdm
            match_data.append(sheet_pdm[('E' + str(i))].value) # revised by pdm
            match_data.append(sheet_arena[('C' + str(x))].value) #revision arena
            match_data.append(sheet_arena[('D' + str(x))].value) #phase state arena
            master_data = numpy.vstack((master_data,match_data))
            # matching_table.append(current_value_last)
            # table_index+=1



#initialize the various combinations of tables":
#Arena and PDM revision match, PDM state is either "Approved (Prototype)" or "Approved (Production)
approved_match = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']

#Arena and PDM revision match, PDM state is "Waiting for approval (initial release)" or "Waiting for approval (Production)"
waiting_match = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']

#Arena and PDM revisions match, PDM state is "Change in Progress (Production)" or "Initial State (ACT)"
change_match = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']

#Arena and PDM revisions do not match
no_match = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']

#Arena and PDM revisions match, PDM state is "ACT Obsolete"
obsolete_match = ['original file name', 'name', 'pdm_revision', 'pdm_state', 'revised by', 'arena_revision', 'item_phase']

#various PDM states
s1="Approved (Production)"
s2="Approved (Prototype)"
s3 = "Waiting for approval (initial release)"
s4 = "Waiting for approval (Production)"
s5 = "Change in Progress (Production)"
s6 = "Initial State (ACT)"
s7 = "ACT Obsolete"

for i in range(1,len(master_data)):
    #check rev match and
    # a = (master_data[i,1])
    # b = (master_data[i, 2])
    # c = (master_data[i, 3])
    # d = (master_data[i, 4])
    # e = (master_data[i, 5])
    if master_data[i,2]==master_data[i,5] and ((master_data[i,3]==s1) or (master_data[i,3]==s2)):
        approved_match = numpy.vstack((approved_match,master_data[i]))
    elif master_data[i,2]==master_data[i,5] and ((master_data[i,3]==s3) or (master_data[i,3]==s4)):
        waiting_match= numpy.vstack((waiting_match,master_data[i]))
    elif master_data[i,2]==master_data[i,5] and ((master_data[i,3]==s5) or (master_data[i,3]==s6)):
        change_match = numpy.vstack((change_match, master_data[i]))
    elif master_data[i,2]!=master_data[i,5]:
        no_match = numpy.vstack((no_match, master_data[i]))
    elif master_data[i,2]==master_data[i,5]:
        obsolete_match = numpy.vstack((obsolete_match,master_data[i]))





new_wb = openpyxl.Workbook()
new_wb.create_sheet(index=0, title="Matching and Approved")
new_wb.create_sheet(index=1, title="Matching, Waiting Approval")
new_wb.create_sheet(index=2, title="Matching, CIP, Initial")
new_wb.create_sheet(index=3, title="Non-matching")
new_wb.create_sheet(index=4, title="ACT Obsolete")




sheet = new_wb["Matching and Approved"]

if len(approved_match)!=7:
    for i in range(1,len(approved_match)):
         sheet['A' + str(i)] = approved_match[i,1]
         sheet['B' + str(i)] = approved_match[i,2]
         sheet['C' + str(i)] = approved_match[i,3]
         sheet['D' + str(i)] = approved_match[i,4]
         sheet['E' + str(i)] = approved_match[i,5]
         sheet['F' + str(i)] = approved_match[i,6]


sheet = new_wb["Matching, Waiting Approval"]
if len(waiting_match)!=7:
    for i in range(1,len(waiting_match)):
         sheet['A' + str(i)] = waiting_match[i,1]
         sheet['B' + str(i)] = waiting_match[i,2]
         sheet['C' + str(i)] = waiting_match[i,3]
         sheet['D' + str(i)] = waiting_match[i,4]
         sheet['E' + str(i)] = waiting_match[i,5]
         sheet['F' + str(i)] = waiting_match[i,6]

sheet = new_wb["Matching, CIP, Initial"]
if len(change_match)!=7:
    for i in range(1,len(change_match)):
         sheet['A' + str(i)] = change_match[i,1]
         sheet['B' + str(i)] = change_match[i,2]
         sheet['C' + str(i)] = change_match[i,3]
         sheet['D' + str(i)] = change_match[i,4]
         sheet['E' + str(i)] = change_match[i,5]
         sheet['F' + str(i)] = change_match[i,6]

sheet = new_wb["Non-matching"]
if len(no_match)!=7:
    for i in range(1,len(no_match)):
         sheet['A' + str(i)] = no_match[i,1]
         sheet['B' + str(i)] = no_match[i,2]
         sheet['C' + str(i)] = no_match[i,3]
         sheet['D' + str(i)] = no_match[i,4]
         sheet['E' + str(i)] = no_match[i,5]
         sheet['F' + str(i)] = no_match[i,6]

sheet = new_wb["ACT Obsolete"]
if len(obsolete_match)!=7:
    for i in range(1,len(obsolete_match)):
         sheet['A' + str(i)] = obsolete_match[i,1]
         sheet['B' + str(i)] = obsolete_match[i,2]
         sheet['C' + str(i)] = obsolete_match[i,3]
         sheet['D' + str(i)] = obsolete_match[i,4]
         sheet['E' + str(i)] = obsolete_match[i,5]
         sheet['F' + str(i)] = obsolete_match[i,6]




new_wb.save('first_excel_output.xlsx')
print(approved_match)
print(waiting_match)